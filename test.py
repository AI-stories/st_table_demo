# ---
# jupyter:
#   jupytext:
#     text_representation:
#       extension: .py
#       format_name: light
#       format_version: '1.5'
#       jupytext_version: 1.16.7
#   kernelspec:
#     display_name: .venv
#     language: python
#     name: python3
# ---

# +
import logging
import os
from datetime import datetime
from typing import Literal

import pandas as pd
from dotenv import load_dotenv
from openai import OpenAI
from pydantic import BaseModel, Field
from rich import print
from tqdm import tqdm

client = OpenAI()

logging.basicConfig(level=logging.WARNING)
load_dotenv()
api_key = os.getenv("OPENAI_API_KEY")
client = OpenAI(api_key=api_key)
df = pd.read_excel("Wedding Values.xlsx")
df.columns

# +
# from wedding_venue_models import (
#     WeddingContactInfo,
#     WeddingPriceInfo,
#     WeddingVenueStyle,
#     WeddingVenueOther,
#     WeddingFoodInfo,
# )

# +
# from pathlib import Path

# path = Path(
#     "/Users/mac-robertsocolewicz/Documents/private/playground_tables/test_md/Almansor Court.md"
# )
# venue_name = path.name.replace(path.suffix, "")

# with open(path) as f:
#     doc = f.read()

# system_prompt = """
# You are a helpful wedding AI assistant. Guide the user through understanding various options and pricing for the following wedding venue:

# ===
# venue name: {venue_name}

# venue description:

# {doc}

# ===
# """

# user_prompt = """
# Please provide all the contact info for this wedding venue.
# """

# completion = client.beta.chat.completions.parse(
#     model="gpt-4o-mini",
#     messages=[
#         {
#             "role": "system",
#             "content": system_prompt,
#         },
#         {"role": "user", "content": user_prompt},
#     ],
#     response_format=WeddingContactInfo,
# )

# # math_reasoning = completion.choices[0].message

# +
# from pathlib import Path

# def generate_field_instructions(model_class: type[BaseModel]) -> str:
#     """Generate field-specific instructions from a Pydantic model's docstrings."""
#     instructions = []
#     for field_name, field_info in model_class.model_fields.items():
#         docstring = field_info.description or "No description provided."
#         instructions.append(f"- {field_name}: {docstring}")
#     return "\n".join(instructions)


# def create_system_prompt(model_class: type[BaseModel]) -> str:
#     """Generate a system prompt based on the model class and its field descriptions."""
#     field_instructions = generate_field_instructions(model_class)
#     return f"""
#         You are an expert in wedding planning. You are extracting structured
#         information about wedding venues.

#         First, carefully analyze all relevant information in the text. Consider
#         both explicit statements and reasonable inferences.

#         Important instructions:
#         1. For each field, follow the specific
#            guidelines below about how to handle ambiguous or missing information.
#         2. For boolean fields, return true/false values rather than "Yes"/"No"
#            strings.
#         3. For string fields, provide detailed information or null if
#            not available.
#         4. For numerical fields, use -1 if information is not
#            available.
#         5. Begin by developing a comprehensive reasoning that
#            considers all evidence before determining individual field values.

#         Field-specific instructions:
#         {field_instructions}
#         """


# # Setup logging and OpenAI client
# logging.basicConfig(level=logging.WARNING)
# load_dotenv()
# api_key = os.getenv("OPENAI_API_KEY")
# client = OpenAI(api_key=api_key)

# # List of all Pydantic models to process
# models = [
#     WeddingContactInfo,
#     WeddingPriceInfo,
#     WeddingVenueStyle,
#     WeddingVenueOther,
#     WeddingFoodInfo,
# ]

# # Process venues
# venue_data = []
# test_md_path = Path("test_md")
# if not test_md_path.exists():
#     print(f"Warning: {test_md_path} directory not found")
#     exit()

# md_files = list(test_md_path.glob("*.md"))
# ai_model = "gpt-4.5-preview"

# for file in tqdm(md_files[0:5], desc="Processing venues", unit="file"):
#     tqdm.write(f"Processing: {file.name}")
#     with open(file, "r", encoding="utf-8") as f:
#         md_content = f.read()

#     venue_name = file.stem
#     venue_dict = {"name": venue_name}

#     # Process each model for this venue
#     for model_class in models:
#         system_prompt = create_system_prompt(model_class)
#         try:
#             completion = client.beta.chat.completions.parse(
#                 model=ai_model,
#                 messages=[
#                     {"role": "system", "content": system_prompt},
#                     {
#                         "role": "user",
#                         "content": f"Extract venue information from this text about '{venue_name}':\n\n{md_content}",
#                     },
#                 ],
#                 response_format=model_class,
#                 temperature=0,
#             )
#             venue_info = completion.choices[0].message.parsed.model_dump()
#             # Prefix keys to avoid collisions between models
#             prefixed_venue_info = {
#                 f"{model_class.__name__}_{k}": v for k, v in venue_info.items()
#             }
#             venue_dict.update(prefixed_venue_info)
#             tqdm.write(
#                 f"✓ Successfully processed {model_class.__name__} for: {venue_name}"
#             )
#         except Exception as e:
#             tqdm.write(
#                 f"✗ Error processing {model_class.__name__} for {venue_name}: {e}"
#             )

#     venue_data.append(venue_dict)

# # Create and save DataFrame
# if venue_data:
#     df = pd.DataFrame(venue_data)
#     print(f"\nProcessed {len(venue_data)} venues")
#     print(df)

#     now = datetime.now().strftime("%Y%m%d_%H%M%S")
#     output_path = f"venues_data_{now}.csv"
#     df["model"] = ai_model
#     df.to_csv(output_path, index=False)
#     print(f"Data saved to {output_path}")
# else:
#     print("No venue data was processed")
# -


# +

TierName = Literal["Standard", "Signature", "Premium"]


class WeddingContactInfo(BaseModel):
    city: str | None = Field(description="The city this wedding venue is located in.")
    state: str | None = Field(description="The state this wedding venue is located in.")
    zip_code: str | None = Field(
        description="The zip code this wedding venue is located in."
    )
    country: str | None = Field(
        description="The country this wedding venue is located in."
    )
    email: str | None = Field(description="The email address of the wedding venue.")
    phone: str | None = Field(description="The phone number of the wedding venue.")
    website: str | None = Field(description="The website of the wedding venue.")


class MenuTier(BaseModel):
    name: TierName = Field(
        description="Tier name: must be Standard, Signature, or Premium."
    )
    appetizers: str | None = Field(description="Appetizers in this tier.")
    entrees: str | None = Field(description="Entrées in this tier.")
    sides: str | None = Field(description="Side dishes in this tier.")
    desserts: str | None = Field(description="Desserts in this tier.")
    beverages: str | None = Field(description="Beverages in this tier.")

    def to_string(self) -> str:
        parts = [f"{self.name}"]
        if self.appetizers:
            parts.append(f"- Appetizers: {self.appetizers}")
        if self.entrees:
            parts.append(f"- Entrées: {self.entrees}")
        if self.sides:
            parts.append(f"- Sides: {self.sides}")
        if self.desserts:
            parts.append(f"- Desserts: {self.desserts}")
        if self.beverages:
            parts.append(f"- Beverages: {self.beverages}")
        if len(parts) == 1:
            return ""
        return "\n".join(parts)


class FoodBreakdown(BaseModel):
    """
    You must return exactly three food tiers: Standard, Signature, and
    Premium.

    Venues often name their packages arbitrarily (e.g., "Adore", "Silver",
    "Classic", "Treasure"). Your task is to interpret these names and map each one
    to a standard tier by evaluating the **menu quality, service level, and
    pricing**.

    Use the following general guidance:

    - **Standard**: basic offerings, lower price point
    - **Signature**: moderate enhancements, mid-range price
    - **Premium**: multiple high-end appetizers, upgraded entrées (beef,
        seafood), top-tier service, highest price

    Ignore the original package names. Focus only on the **content and price range**
    to assign the correct standard tier name.
    """

    tiers: list[MenuTier] = Field(
        description="List of exactly 3 tiers: Standard (Silver), Signature (Gold), Premium (Platinum)."
    )
    flexibility: Literal[
        "0: Completely fixed packages, no flexibility",
        "1: Fixed packages with a few extras or options",
        "2: Moderate or flexible approach",
        "3: Highly customizable with some structure",
        "4: Completely custom/DIY",
    ] = Field(
        description="How much freedom does the customer have to customize the food-related tier packages?"
    )

    def to_string(self) -> str:
        return "\n\n".join([tier.to_string() for tier in self.tiers])


class BarTier(BaseModel):
    name: TierName = Field(
        description="Tier name: must be Standard, Signature, or Premium."
    )
    highlights: str | None = Field(
        description="Beverages and package highlights in this drink package."
    )

    bar_pricing_model: Literal["Open bar", "Hosted bar", "Cash bar", "Not Offered"] = (
        Field(
            description="""- open bar: the venue provides the bar and the drinks, prepaid for by the host
- hosted bar: the venue provides the bar and the drinks, paid for by the host at the end of the night, also referred to as a consumption bar or tab bar
- cash bar: the venue provides the bar, but the drinks are payed for by the guests
- not offered: the venue does not offer a bar or is not specified"""
        )
    )

    def to_string(self) -> str:
        parts = [f"{self.name}"]
        if self.highlights:
            parts.append(f"- Highlights: {self.highlights}")
        if self.bar_pricing_model != "Not Offered":
            parts.append(f"- Bar Pricing Model: {self.bar_pricing_model}")
        return "\n".join(parts)


class BarBreakdown(BaseModel):
    """
    You must return exactly three bar tiers: Standard, Signature, and
    Premium.

    Many venues use unique names for their drink packages. Instead of
    copying those names, evaluate the **included alcohol types and
    service level**, and map them to a standard tier:

    - **Standard**: basic offerings, lower price point
    - **Signature**: mid-range price, includes house liquors or a soft
        bar with wine service
    - **Premium**: top-tier service, highest price like top-shelf
        liquor, signature cocktails, champagne toast, or full open bar

    Always normalize to these three tiers based on the drink offerings —
    not the label.
    """

    tiers: list[BarTier] = Field(
        description="List of exactly 3 tiers: Standard (Silver), Signature (Gold), Premium (Platinum)."
    )
    flexibility: Literal[
        "0: Completely fixed packages, no flexibility",
        "1: Fixed packages with a few extras or options",
        "2: Moderate or flexible approach",
        "3: Highly customizable with some structure",
        "4: Completely custom/DIY",
    ] = Field(
        description="How much freedom does the customer have to customize the package?"
    )

    def to_string(self) -> str:
        parts = []
        for tier in self.tiers:
            parts.append(tier.to_string())

        return "\n".join(parts)


class WeddingPriceInfo(BaseModel):
    option: Literal["standard", "premium", "signature"] = Field(
        description="This is the pricing option for this wedding venue."
    )
    ceremony_cost: int | Literal["unknown"] = Field(
        description="Estimate a cost for the ceremony. Respond with a number only which would be your best guess based on the available information. Use 'unknown' if no information is provided."
    )

    def to_string(self) -> str:
        if self.ceremony_cost:
            return f"Ceremony: {self.option}\n - ceremony cost: ${self.ceremony_cost}\n"
        return ""


class WeddingVenuePricingSummary(BaseModel):
    options: list[WeddingPriceInfo] = Field(
        description="This is a list of all the pricing options for this wedding venue."
    )
    price: int = Field(
        description="""Analyze this wedding venue document and calculate the PER PERSON cost in USD for a 100-guest wedding reception.

Important instructions:
1. Return ONLY a single number representing the per-person cost in USD (e.g., 250)
2. Do NOT return the total venue cost - I need the PER PERSON amount
3. If you see a total price (e.g., $20,000 for 80 guests), divide by the number of guests to get the per-person rate ($20,000 ÷ 80 = $250 per person)
4. Then adjust this per-person rate for a 100-person event, accounting for any tiered pricing

Your calculation must include:
- Food (dinner, appetizer, dessert)
- Beverage/alcohol package
- Venue rental fees (divided by 100 guests)
- All taxes, service charges, and gratuities

Exclude:
- Ceremony costs (focus only on reception)
- Photography
- Late night food
- Any optional add-ons

If multiple packages exist, use the middle-tier or "standard" option.
If specific costs aren't provided, estimate based on the venue's location and quality level.

Think through your calculation step by step:
1. Identify the base per-person food & beverage cost
2. Add any flat venue fees (calculate the per-person amount)
3. Add per-person share of taxes and service charges
4. Sum these components for the final per-person total

Remember: Your answer must be ONLY a single number representing the per-person cost in USD."""
    )
    base_prices: str = Field(
        description="The base cost of the venue, including menu items, beverages and any variations based on day of the week or package. Break down ALL venue costs on a per-person basis. For example, if the venue has a $20,000 base price for 80 guests, the base price per person is $250."
    )
    taxes_and_fees: str = Field(
        description="The breakdown of taxes, gratuity, and any additional fees, including their individual amounts and the total. Convert ALL taxes, service charges, and fees to per-person amounts. For example, if the venue has a $2,185 tax for 80 guests, the tax per person is $27.31."
    )
    pricing_transparency: Literal[
        "1: This venue discloses a small portion of the total wedding costs",
        "2: This venue discloses a moderate portion of the total wedding costs",
        "3: This venue discloses a high degree of the total costs",
        "X: Not enough information",
    ] = Field(
        description="""Assess how much of the total wedding cost is disclosed in the provided materials. You MUST CHOOSE ONE of the following options that best matches the document's content and return the selected option's description as the field value:

- This venue discloses a small portion of the total wedding costs
- This venue discloses a moderate portion of the total wedding costs
- This venue discloses a high degree of the total wedding costs
- Not enough information

Guidance when selecting the option: A 'small portion' of disclosure means significant costs (e.g., food, bar/alcohol) are unclear or require contacting external vendors. A 'moderate portion' means some unknowns exist, but you can get a general cost idea without much extra work. A 'high degree' means most costs are disclosed with few surprises, little additional work needed to understand the total cost."""
    )
    flexibility: Literal[
        "0: Completely fixed packages, no flexibility",
        "1: Fixed packages with a few extras or options",
        "2: Moderate or flexible approach",
        "3: Highly customizable with some structure",
        "4: Completely custom/DIY",
    ] = Field(
        description="How much freedom does the customer have to customize the wedding/ceremony package?"
    )
    deposit_and_payment_plans: Literal[
        "The venue works with me on deposit terms and payment plans",
        "The venue does not have flexibility on deposit terms and payment plans",
        "Not enough information",
    ] = Field(
        description="""Determine if the venue offers flexibility on deposit terms and payment plans. You MUST CHOOSE ONE of the following options that best matches the document's content and return the selected option's description as the field value:

- The venue works with me on deposit terms and payment plans
- The venue does not have flexibility on deposit terms and payment plans
- Not enough information.

Follow these guidance when selecting the option: Flexibility means the venue allows negotiation on deposit amounts, payment schedules, or offers installment plans. Lack of flexibility is indicated by strict terms or no mention of flexible options."""
    )

    def to_string(self) -> str:
        parts = []
        if self.options:
            for option in self.options:
                parts.append(option.to_string())
        return "\n".join(parts)


class WeddingVenueStyle(BaseModel):
    style: Literal[
        "Barns & Farms",
        "Hotels",
        "Winery",
        "Country Clubs",
        "Restaurants",
        "Rooftops & Lofts",
        "Mansions",
        "Religious Spaces",
        "Museums",
        "Boats",
        "Parks",
        "Historic Venues",
        "Banquet Halls",
        "Beach",
        "Garden",
        "Waterfront",
        "Brewery",
        "State",
        "Local",
        "Government Property",
        "Other",
    ] = Field(
        description="""Determine the style of the venue. You MUST CHOOSE ONE of the following options that best matches the document's content and return the selected option's value as the field value:
- Barns & Farms
- Hotels
- Winery
- Country Clubs
- Restaurants
- Rooftops & Lofts
- Mansions
- Religious Spaces
- Museums
- Boats
- Parks
- Historic Venues
- Banquet Halls
- Beach
- Garden
- Waterfront
- Brewery
- State
- Local
- Government Property
- Other"""
    )
    indoor_outdoor: Literal[
        "1: Completely indoor",
        "2: Predominantly indoor",
        "3: Equal indoor/outdoor mix",
        "4: Predominantly outdoor",
        "5: Completely outdoor",
        "X: Not enough information",
    ] = Field(
        description="""Determine the indoor/outdoor nature of the venue. You MUST CHOOSE ONE of the following options and return the selected option's description as the field value:
- "1: Completely indoor"
- "2: Predominantly indoor"
- "3: Equal indoor/outdoor mix"
- "4: Predominantly outdoor"
- "5: Completely outdoor"
- "X: Not enough information"

For guidance: You must choose one of the above options. A predominantly outdoor space may include gardens or patios, while one without a covering lacks weather protection (e.g., an open field)."""
    )
    privacy: Literal[
        "1: Privacy and exclusivity is a major feature of the venue",
        "2: Moderate privacy with possible nearby non-wedding guests",
        "3: Shared or public space",
        "X: Not enough information",
    ] = Field(
        description="""Assess the privacy level. You MUST CHOOSE ONE of the following options and return the selected option's description as the field value:
- "1: Privacy and exclusivity is a major feature of the venue"
- "2: Moderate privacy with possible nearby non-wedding guests"
- "3: Shared or public space"
- "X: Not enough information"."""
    )
    accommodations: Literal[
        "1: On-site lodging accommodations or extremely close (less than 5 mins walking distance) is possible",
        "2: No On-site lodging, third-party 5-10 minutes away.",
        "3: No on-site lodging, third-party 10+ minutes away.",
        "X: Not enough information",
    ] = Field(
        description="""Determine the lodging accommodations offered by the venue. You MUST CHOOSE ONE of the following options that best matches the document's content and return the selected option's description as the field value:
- "1: On-site lodging accommodations or extremely close (less than 5 mins walking distance) is possible"
- "2: No On-site lodging, third-party 5-10 minutes away."
- "3: No on-site lodging, third-party 10+ minutes away."
- "X: Not enough information"."""
    )
    environmental: Literal[
        "This venue focuses on minimal environmental impact and sustainability in their offering",
        "This venue does not emphasize environment or sustainability",
    ] = Field(
        description="""Assess the venue's emphasis on environmental sustainability. You MUST CHOOSE ONE of the following options that best matches the document's content and return the selected option's description as the field value:
- This venue focuses on minimal environmental impact and sustainability in their offering
- This venue does not emphasize environment or sustainability.

For guidance: A focus on sustainability might include eco-friendly practices (e.g., local sourcing or renewable energy), while lack of emphasis implies no mention of such efforts."""
    )
    general_vibe: Literal[
        "Rustic and simple",
        "Peaceful and serene",
        "Grandiose and elegant",
        "Adventurous or quirky",
        "Warm and cozy",
        "Other",
    ] = Field(
        description="""Determine the general vibe or atmosphere of the venue. You MUST CHOOSE ONE of the following options that best matches the document's content and return the selected option's description as the field value:
- Rustic and simple
- Peaceful and serene
- Grandiose and elegant
- Adventurous or quirky
- Warm and cozy
- Other. For guidance: Rustic and simple might include barns or farms, peaceful and serene could be gardens, grandiose and elegant might be mansions, adventurous or quirky could involve unique spaces, and warm and cozy might be cozy interiors."""
    )


class WeddingVenueOther(BaseModel):
    outside_wedding_coordinator: Literal[True, False, "unknown"] = Field(
        description="This venue lets me bring in my own wedding coordinator"
    )
    outside_photographer: Literal[True, False, "unknown"] = Field(
        description="This venue lets me bring in my own photographer"
    )
    package_approach: Literal[
        "1: This venue offers fixed packages, with a few extras or options",
        "2: This venue offers a moderate or flexible approach",
        "3: This venue provides a high degree of flexibility",
        "X: Not enough information",
    ] = Field(
        description="""Determine the wedding package approach of the venue. You MUST CHOOSE ONE of the following options that best matches the document's content and return the selected option's description as the field value:
- "1: This venue offers fixed packages, with a few extras or options"
- "2: This venue offers a moderate or flexible approach"
- "3: This venue provides a high degree of flexibility"
- "X: Not enough information"

For guidance: Fixed packages include in-house services (e.g., pre-set menus); moderate offers a mix; high flexibility requires client planning."""
    )
    reception_or_ceremony: Literal[
        "1: Only space for reception",
        "2: Only space for the ceremony",
        "3: Space for both reception and the ceremony",
        "X: Not enough information",
    ] = Field(
        description="""Determine an option for spaces provided by the venue. You MUST CHOOSE ONE of the following options that best matches the document's content and return the selected option's description as the field value:
- "1: Only space for reception"
- "2: Only space for the ceremony"
- "3: Space for both reception and the ceremony"
- "X: Not enough information"."""
    )
    what_time_does_the_party_need_to_stop: Literal[
        "1: 10PM",
        "2: 12AM",
        "3: After 12AM",
        "X: Not enough information",
    ] = Field(
        description="""Determine the latest time the event must conclude. You MUST CHOOSE ONE of the following options that best matches the document's content:
- "1: 10PM"
- "2: 12AM"
- "3: After 12AM"
- "X: Not enough information"."""
    )
    top_choices: str = Field(
        description="Give me 1-2 reasons why you think a client is most likely to choose this venue (e.g., what are the biggest key selling points or benefits)."
    )
    guest_capacity: Literal[1, 50, 100, 150, 200, 300] = Field(
        description="""Determine the estimated guest capacity. You MUST CHOOSE ONE of the following approximate maximums that best matches the document's content, representing the upper bound of each range:
- 1 (for 1-50)
- 50 (for 50-100)
- 100 (for 100-150)
- 150 (for 150-200)
- 200 (for 200-300)
- 300 (for 300+)
Guidance: Infer from max guests or room size (e.g., 'up to 150 guests' = 150)."""
    )


def generate_field_instructions(model_class: type[BaseModel]) -> str:
    """Generate field-specific instructions from a Pydantic model's docstrings."""
    instructions = []
    for field_name, field_info in model_class.model_fields.items():
        docstring = (
            field_info.description or field_info.__doc__
        ) or "No description provided."
        instructions.append(f"- {field_name}: {docstring}")
    return "\n".join(instructions)


def create_system_prompt(model_class: type[BaseModel]) -> str:
    field_instructions = generate_field_instructions(model_class)
    tier_hint = model_class.__doc__

    return f"""
        You are an expert in wedding planning. You are extracting structured
        information about wedding venues.

        First, carefully analyze all relevant information in the text. Consider
        both explicit statements and reasonable inferences.

        Important instructions: 
        1. For each field, follow the specific guidelines below about how to handle 
           ambiguous or missing information.
        2. For boolean fields, return true/false values rather than "Yes"/"No" strings. 
        3. For string fields, provide detailed information or null if not available. 
        4. Begin by developing a comprehensive reasoning that considers all evidence 
           before determining individual field values.

        Field-specific instructions:
        {tier_hint}
        {field_instructions}
        """


class WeddingFoodInfo(BaseModel):
    outside_food_allowed: Literal[True, False] = Field(
        description="""Can external catering or food vendors be used at this venue?If the text explicitly mentions allowing outside catering or food vendors, answer True. Do not consider desserts or cakes as food, as these are different categories. If not mentioned, default to False."""
    )
    outside_alcohol_allowed: Literal[True, False] = Field(
        description="""Can external alcohol be brought into this venue?If the text mentions the venue has a liquor license or provides alcohol service, answer False. If the text explicitly mentions allowing outside alcohol or BYOB policy, answer True. If not mentioned, default to False."""
    )
    outside_dessert_allowed: Literal[True, False] = Field(
        description="""Can external desserts or cakes be brought into this venue?If the text explicitly mentions allowing outside desserts or cakes, answer True. If not mentioned, default to True."""
    )
    kosher_food: Literal[True, False] = Field(
        description="""Does the venue offer certified kosher menu options, either in-house or through external vendors?Answer True ONLY if kosher options are explicitly mentioned or if the venue explicitly states they accommodate religious dietary restrictions. If not mentioned, default to False."""
    )
    halal_food: Literal[True, False] = Field(
        description="""Does the venue offer certified halal menu options, either in-house or through external vendors? Return True if offered, False if not offered. Answer True ONLY if halal options are explicitly mentioned. If not mentioned, default to False."""
    )
    east_asian_food: Literal[True, False] = Field(
        description="""Does the venue offer East Asian food options (Chinese, Japanese, Korean, etc.), either in-house or through external vendors? Answer True if East Asian cuisine is explicitly mentioned or if the venue offers international cuisine. If not mentioned, default to False."""
    )
    indian_food: Literal[True, False] = Field(
        description="""Does the venue offer Indian food options, either in-house or through external vendors? Answer True if Indian cuisine is explicitly mentioned or if the venue offers diverse international cuisine that likely includes Indian options. If not mentioned, default to False."""
    )
    gluten_free_food: Literal[True, False] = Field(
        description="""Does the venue offer gluten-free food options? Return True if offered, False if not offered. Answer True if gluten-free options are explicitly mentioned. If the venue mentions accommodating dietary restrictions/allergies/food preferences, return True."""
    )
    other_ethnic_food_style: str | Literal["unknown"] = Field(
        description="""Does the venue offer other ethnic food styles beyond those already mentioned? If so, list the available styles as a comma-separated string. If no other ethnic food styles are mentioned, return 'unknown'."""
    )
    late_night_food: Literal[True, False] = Field(
        description="""Does the venue provide late-night food options, such as pizza or snacks after dinner and dessert service? If not mentioned, default to False."""
    )


# -


class WeddingVenuePricingSummary(BaseModel):
    options: list[WeddingPriceInfo] = Field(
        description="This is a list of all the pricing options for this wedding venue."
    )
    price: int = Field(
        description="""Analyze this wedding venue document and calculate the PER PERSON cost in USD for a 100-guest wedding reception.

Important instructions:
1. Return ONLY a single number representing the per-person cost in USD (e.g., 250)
2. Do NOT return the total venue cost - I need the PER PERSON amount
3. If you see a total price (e.g., $20,000 for 80 guests), divide by the number of guests to get the per-person rate ($20,000 ÷ 80 = $250 per person)
4. Then adjust this per-person rate for a 100-person event, accounting for any tiered pricing

Your calculation must include:
- Food (dinner, appetizer, dessert)
- Beverage/alcohol package
- Venue rental fees (divided by 100 guests)
- All taxes, service charges, and gratuities

Exclude:
- Ceremony costs (focus only on reception)
- Photography
- Late night food
- Any optional add-ons

If multiple packages exist, use the middle-tier or "standard" option.
If specific costs aren't provided, estimate based on the venue's location and quality level.

Think through your calculation step by step:
1. Identify the base per-person food & beverage cost
2. Add any flat venue fees (calculate the per-person amount)
3. Add per-person share of taxes and service charges
4. Sum these components for the final per-person total

Remember: Your answer must be ONLY a single number representing the per-person cost in USD."""
    )
    base_prices: str = Field(
        description="The base cost of the venue, including menu items, beverages and any variations based on day of the week or package. Break down ALL venue costs on a per-person basis. For example, if the venue has a $20,000 base price for 80 guests, the base price per person is $250."
    )
    taxes_and_fees: str = Field(
        description="The breakdown of taxes, gratuity, and any additional fees, including their individual amounts and the total. Convert ALL taxes, service charges, and fees to per-person amounts. For example, if the venue has a $2,185 tax for 80 guests, the tax per person is $27.31."
    )
    pricing_transparency: Literal[
        "1: This venue discloses a small portion of the total wedding costs",
        "2: This venue discloses a moderate portion of the total wedding costs",
        "3: This venue discloses a high degree of the total costs",
        "X: Not enough information",
    ] = Field(
        description="""Assess how much of the total wedding cost is disclosed in the provided materials. You MUST CHOOSE ONE of the following options that best matches the document's content and return the selected option's description as the field value:

- This venue discloses a small portion of the total wedding costs
- This venue discloses a moderate portion of the total wedding costs
- This venue discloses a high degree of the total wedding costs
- Not enough information

Guidance when selecting the option: A 'small portion' of disclosure means significant costs (e.g., food, bar/alcohol) are unclear or require contacting external vendors. A 'moderate portion' means some unknowns exist, but you can get a general cost idea without much extra work. A 'high degree' means most costs are disclosed with few surprises, little additional work needed to understand the total cost."""
    )
    flexibility: Literal[
        "0: Completely fixed packages, no flexibility",
        "1: Fixed packages with a few extras or options",
        "2: Moderate or flexible approach",
        "3: Highly customizable with some structure",
        "4: Completely custom/DIY",
    ] = Field(
        description="How much freedom does the customer have to customize the wedding/ceremony package?"
    )
    deposit_and_payment_plans: Literal[
        "The venue works with me on deposit terms and payment plans",
        "The venue does not have flexibility on deposit terms and payment plans",
        "Not enough information",
    ] = Field(
        description="""Determine if the venue offers flexibility on deposit terms and payment plans. You MUST CHOOSE ONE of the following options that best matches the document's content and return the selected option's description as the field value:

- The venue works with me on deposit terms and payment plans
- The venue does not have flexibility on deposit terms and payment plans
- Not enough information.

Follow these guidance when selecting the option: Flexibility means the venue allows negotiation on deposit amounts, payment schedules, or offers installment plans. Lack of flexibility is indicated by strict terms or no mention of flexible options."""
    )

    def to_string(self) -> str:
        parts = []
        if self.options:
            for option in self.options:
                parts.append(option.to_string())
        return "\n".join(parts)


WeddingContactInfo.model_fields["city"]

print(system_prompt)

# +
from pathlib import Path

venue_data = []
md_path = Path("test_md")
if not md_path.exists():
    raise FileNotFoundError("Directory 'test_md' not found")

md_files = list(md_path.glob("*.md"))[-1:]
md_files = [Path("test_md/a.o.c. Brentwood.md")]
venue_data = None
for file in tqdm(md_files, desc="Processing venues", unit="file"):
    tqdm.write(f"Processing: {file.name}")
    with open(file, "r", encoding="utf-8") as f:
        md_content = f.read()

    raw = []
    venue_name = file.stem

    system_prompt = create_system_prompt(WeddingVenuePricingSummary)

completion = client.beta.chat.completions.parse(
    model="o3-mini",
    messages=[
        {"role": "system", "content": system_prompt},
        {
            "role": "user",
            "content": f"Extract venue information from this text about '{venue_name}':\n\n{md_content}",
        },
    ],
    response_format=WeddingVenuePricingSummary,
)
x = completion.choices[0].message.parsed
print("price: ", x.price)
print("base_prices: ", x.base_prices)
print("taxes_and_fees: ", x.taxes_and_fees)

# -

models = [
    WeddingContactInfo,
    FoodBreakdown,
    WeddingFoodInfo,
    BarBreakdown,
    WeddingVenuePricingSummary,
    WeddingVenueStyle,
    WeddingVenueOther,
]

# +
from itertools import chain
from typing import get_args

import numpy as np
from openpyxl.styles import Font, PatternFill


def assert_keys_in_readable_columns(
    models: list[type[BaseModel]], readable_columns: dict[str, str]
) -> None:
    keys = list(
        chain.from_iterable(
            [
                model.__name__ + "_" + np.array(list(model.model_fields.keys()))
                for model in models
            ]
        )
    )
    keys = [
        str(key).replace("_tiers", "_summary").replace("_options", "_summary")
        for key in keys
    ]
    keys
    assert set(keys) - set(readable_columns.keys()) == set(), (
        f"missing keys in readable_columns: {set(keys) - set(readable_columns.keys())}"
    )


readable_columns = {
    "venue": "wedding venue",
    "WeddingVenuePricingSummary_price": "price per guest",
    "WeddingVenuePricingSummary_base_prices": "price breakdown",
    "WeddingVenuePricingSummary_taxes_and_fees": "price breakdown taxes and fees",
    "WeddingVenuePricingSummary_flexibility": "venue customization flexibility",
    # "WeddingPriceInfo_option": "options",
    "WeddingContactInfo_city": "city",
    "WeddingContactInfo_state": "state",
    "WeddingContactInfo_country": "country",
    "WeddingContactInfo_zip_code": "zip code",
    "WeddingContactInfo_email": "email",
    "WeddingContactInfo_website": "website",
    "WeddingContactInfo_phone": "phone",
    "WeddingVenuePricingSummary_summary": "venue pricing summary",
    "FoodBreakdown_summary": "food menu breakdown",
    "FoodBreakdown_flexibility": "food menu flexibility",
    "BarBreakdown_summary": "bar menu breakdown",
    "BarBreakdown_flexibility": "bar menu flexibility",
    "WeddingVenuePricingSummary_pricing_transparency": "pricing transparency",
    "WeddingVenuePricingSummary_deposit_and_payment_plans": "deposit and payment plans",
    "WeddingVenueStyle_style": "style",
    "WeddingVenueStyle_indoor_outdoor": "indoor/outdoor seating",
    "WeddingVenueStyle_privacy": "privacy",
    "WeddingVenueStyle_accommodations": "accommodations",
    "WeddingVenueStyle_environmental": "environmental",
    "WeddingVenueStyle_general_vibe": "general vibe",
    "WeddingFoodInfo_east_asian_food": "serves east asian food",
    "WeddingFoodInfo_gluten_free_food": "serves gluten free food",
    "WeddingFoodInfo_halal_food": "serves halal food",
    "WeddingFoodInfo_indian_food": "serves indian food",
    "WeddingFoodInfo_kosher_food": "serves kosher food",
    "WeddingFoodInfo_late_night_food": "serves late night food",
    "WeddingFoodInfo_other_ethnic_food_style": "serves other ethnic food",
    "WeddingFoodInfo_outside_alcohol_allowed": "allows outside alcohol",
    "WeddingFoodInfo_outside_dessert_allowed": "allows outside dessert",
    "WeddingFoodInfo_outside_food_allowed": "allows outside food",
    "WeddingVenueOther_guest_capacity": "guest capacity",
    "WeddingVenueOther_what_time_does_the_party_need_to_stop": "what time does the party need to stop",
    "WeddingVenueOther_outside_photographer": "allows outside photographer",
    "WeddingVenueOther_package_approach": "package approach",
    "WeddingVenueOther_outside_wedding_coordinator": "allows outside wedding coordinator",
    "WeddingVenueOther_reception_or_ceremony": "reception or ceremony",
    "WeddingVenueOther_top_choices": "top choices",
}


assert_keys_in_readable_columns(models, readable_columns)


def flatten_dict(d: dict, parent_key: str = "", sep: str = "_") -> dict:
    """Flatten a nested dictionary by concatenating nested keys with a separator.

    Parameters
    ----------
    d : dict
        The dictionary to flatten
    parent_key : str, optional
        The parent key for nested dictionaries, by default ""
    sep : str, optional
        The separator to use between nested keys, by default "_"

    Returns
    -------
    dict
        A flattened dictionary with concatenated keys

    Examples
    --------
    >>> d = {"a": 1, "b": {"c": 2, "d": {"e": 3}}}
    >>> flatten_dict(d)
    {'a': 1, 'b_c': 2, 'b_d_e': 3}
    """
    items: list = []
    for k, v in d.items():
        new_key = f"{parent_key}{sep}{k}" if parent_key else k
        if isinstance(v, dict):
            items.extend(flatten_dict(v, new_key, sep).items())
        else:
            items.append((new_key, v))
    return dict(items)


class WeddingVenue:
    def __init__(self, venue_name: str, raw: list[BaseModel]):
        item_dict = {"venue": venue_name}
        for item in raw:
            obj_dict = item.model_dump()
            if "tiers" in obj_dict:
                obj_dict.pop("tiers")
                obj_dict["summary"] = item.to_string()

            if "options" in obj_dict:
                obj_dict.pop("options")
                obj_dict["summary"] = item.to_string()

            item_dict[item.__class__.__name__] = obj_dict

        self.df = pd.DataFrame()
        self.update(item_dict)

    def add_price_breakdown(self) -> None:
        self.df["price breakdown"] = self.df[
            [
                "price breakdown",
                "price breakdown taxes and fees",
            ]
        ].apply(
            lambda x: f"""
                base prices: {x.iloc[0]}
                taxes and fees: {x.iloc[1]}
                """,
            axis=1,
        )
        del self.df["price breakdown taxes and fees"]

    def add_bar_flexibility(self) -> None:
        self.df["bar menu flexibility info"] = self.df["bar menu flexibility"]
        args = get_args(BarBreakdown.model_fields["flexibility"].annotation)
        self.df["bar menu flexibility"] = self.df["bar menu flexibility"].map(
            lambda x: len(args) - args.index(x)
        )

    def add_indoor_outdoor_seating(self) -> None:
        self.df["indoor/outdoor seating info"] = self.df["indoor/outdoor seating"]
        args = get_args(WeddingVenueStyle.model_fields["indoor_outdoor"].annotation)
        self.df["indoor/outdoor seating"] = self.df["indoor/outdoor seating"].map(
            lambda x: args.index(x) if args.index(x) != len(args) else "X"
        )

    def update(self, d: dict) -> None:
        self.df = pd.DataFrame(flatten_dict(d), index=[0])
        self.rename_columns()
        self.add_price_breakdown()
        self.add_bar_flexibility()
        self.add_indoor_outdoor_seating()

    def _repr_html_(self) -> str:
        return self.df._repr_html_()

    def rename_columns(self) -> None:
        """Rename and reorder columns based on readable_columns dictionary."""
        self.df.rename(columns=readable_columns, inplace=True)
        self.df.set_index("wedding venue", inplace=True)
        ordered_columns = [
            col for col in readable_columns.values() if col != "wedding venue"
        ]
        self.df = self.df.reindex(columns=ordered_columns)

    def to_excel(self, name: str = "wedding_venue.xlsx"):
        if not name.endswith(".xlsx"):
            name = f"{name}.xlsx"
        if os.path.exists(name):
            name = f"{name}_{datetime.now().strftime('%Y%m%d_%H%M%S')}.xlsx"
        print("saving to: ", name)
        with pd.ExcelWriter(name, engine="openpyxl") as writer:
            self.df.to_excel(writer, sheet_name="Venue Options")

            worksheet = writer.sheets["Venue Options"]

            header_fill = PatternFill(
                start_color="B3E5FC", end_color="B3E5FC", fill_type="solid"
            )
            header_font = Font(bold=True)

            for col in range(1, len(df.columns) + 1):
                cell = worksheet.cell(row=1, column=col)
                cell.fill = header_fill
                cell.font = header_font

            for col in worksheet.columns:
                max_length = 0
                column = col[0].column_letter
                for cell in col:
                    if cell.value:
                        max_length = max(max_length, len(str(cell.value)))
                adjusted_width = max_length + 2
                worksheet.column_dimensions[column].width = min(adjusted_width, 50)

            worksheet.auto_filter.ref = worksheet.dimensions

        self.df.to_excel(writer, sheet_name="Venue Options")
        return self

    def __add__(self, other: "WeddingVenue") -> "WeddingVenue":
        self.df = pd.concat([self.df, other.df])
        return self


# +
# Setup
from pathlib import Path

import openai

load_dotenv()
api_key = os.getenv("OPENAI_API_KEY")
client = OpenAI(api_key=api_key)

venue_data = []
md_path = Path("test_md")
if not md_path.exists():
    raise FileNotFoundError("Directory 'test_md' not found")

md_files = list(md_path.glob("*.md"))[-8:]

venue_data = None
for file in tqdm(md_files, desc="Processing venues", unit="file"):
    tqdm.write(f"Processing: {file.name}")
    with open(file, "r", encoding="utf-8") as f:
        md_content = f.read()

    raw = []
    venue_name = file.stem
    venue_dict = {"name": venue_name}

    for model_class in models:
        system_prompt = create_system_prompt(model_class)
        try:
            if model_class == WeddingVenuePricingSummary:
                ai_model = "o3-mini"
                temperature = openai.NOT_GIVEN
            else:
                ai_model = "gpt-4o-mini"
                temperature = 0
            completion = client.beta.chat.completions.parse(
                model=ai_model,
                messages=[
                    {"role": "system", "content": system_prompt},
                    {
                        "role": "user",
                        "content": f"Extract venue information from this text about '{venue_name}':\n\n{md_content}",
                    },
                ],
                response_format=model_class,
                temperature=temperature,
            )
            obj = completion.choices[0].message.parsed
            raw.append(obj)
            if hasattr(obj, "to_string"):
                string_summary = obj.to_string()
                venue_dict[f"{model_class.__name__}_summary"] = string_summary
            else:
                venue_dict[f"{model_class.__name__}_summary"] = obj.model_dump()

            tqdm.write(f"✓ Processed {model_class.__name__} for: {venue_name}")

        except Exception as e:
            tqdm.write(f"✗ Error with {model_class.__name__} for {venue_name}: {e}")
            venue_dict[f"{model_class.__name__}_summary"] = None
    if venue_data is None:
        venue_data = WeddingVenue(venue_name, raw)
    else:
        try:
            venue_data += WeddingVenue(venue_name, raw)
        except Exception as e:
            print(f"✗ Error adding {venue_name}: {e}")
            # venue_data = None

if venue_data is not None:
    venue_data.to_excel("all_info.xlsx")
else:
    print("⚠️ No venue data processed.")
# -

venue_data + venue_data

WeddingVenue(venue_name, raw) + venue_data

venue = WeddingVenue(venue_name, raw)
venue.to_excel()

raw

# +
# readable_columns = {
#     "venue": "wedding venue",
#     "FoodBreakdown_summary": "food breakdown",
#     "BarBreakdown_flexibility": "bar flexibility",
#     "BarBreakdown_summary": "bar breakdown",
#     "WeddingContactInfo_city": "city",
#     "WeddingContactInfo_state": "state",
#     "WeddingContactInfo_zip_code": "zip code",
#     "WeddingContactInfo_country": "country",
#     "WeddingContactInfo_email": "email",
#     "WeddingContactInfo_phone": "phone",
#     "WeddingContactInfo_website": "website",
#     "WeddingVenuePricingSummary_pricing_transparency": "pricing transparency",
#     "WeddingVenuePricingSummary_deposit_and_payment_plans": "deposit and payment plans",
#     "WeddingVenuePricingSummary_summary": "venue pricing summary",
#     "WeddingVenueStyle_style": "style",
#     "WeddingVenueStyle_indoor_outdoor": "indoor/outdoor",
#     "WeddingVenueStyle_privacy": "privacy",
#     "WeddingVenueStyle_accommodations": "accommodations",
#     "WeddingVenueStyle_environmental": "environmental",
#     "WeddingVenueStyle_general_vibe": "general vibe",
# }

# +
# df.columns

# +
# print(df["FoodBreakdown_summary"].iloc[0])

# +
# # List of all Pydantic models to process
# models = [
#     WeddingContactInfo,
#     WeddingPriceInfo,
#     WeddingVenueStyle,
#     WeddingVenueOther,
#     WeddingFoodInfo,
# ]
# -


# +
# import json
# import logging
# import os
# from datetime import datetime
# from pathlib import Path

# import google.generativeai as genai
# import pandas as pd
# from dotenv import load_dotenv
# from tqdm import tqdm

# # Setup detailed logging
# logging.basicConfig(
#     level=logging.DEBUG, format="%(asctime)s - %(levelname)s - %(message)s"
# )

# # Load environment variables
# load_dotenv()
# # # api_key = os.getenv("GOOGLE_API_KEY")
# # if not api_key:
# #     raise ValueError(
# #         "GOOGLE_API_KEY not found in environment variables. Please set it in your .env file."
# #     )

# # Configure the Gemini API
# api_key = "AIzaSyAnFuVtRCMOC7NMJHgAbfrW9wzDahMG6gY"
# genai.configure(api_key=api_key)


# # Define helper functions
# def generate_field_instructions(model_class: type[BaseModel]) -> str:
#     """Generate field-specific instructions from a Pydantic model's docstrings."""
#     instructions = []
#     for field_name, field_info in model_class.model_fields.items():
#         docstring = field_info.description or "No description provided."
#         # Add type hint to prompt for numeric fields
#         if field_name == "guest_capacity":
#             instructions.append(
#                 f"- {field_name}: {docstring} (Return as an integer: 1, 50, 100, 150, 200, or 300, or -1 if not available)"
#             )
#         else:
#             instructions.append(f"- {field_name}: {docstring}")
#     return "\n".join(instructions)


# def create_system_prompt(model_class: type[BaseModel]) -> str:
#     """Generate a comprehensive prompt based on the model class and its field descriptions."""
#     field_instructions = generate_field_instructions(model_class)
#     return f"""You are an expert in wedding planning. You are extracting structured information about wedding venues.

# First, carefully analyze all relevant information in the text. Consider both explicit statements and reasonable inferences.

# Important instructions:
# 1. For each field, follow the specific guidelines below about how to handle ambiguous or missing information.
# 2. For boolean fields, return true/false values rather than "Yes"/"No" strings.
# 3. For string fields, provide detailed information or null if not available.
# 4. For numerical fields (e.g., guest_capacity), return integers or -1 if not available.
# 5. Begin by developing a comprehensive reasoning that considers all evidence before determining individual field values.
# 6. For fields with predefined choices (e.g., Literal types), you MUST return only the exact values listed in the options. If the information does not match any option, default to 'Not enough information' or the specified default value.
# 7. Return a valid JSON object with no additional text or commentary.

# Field-specific instructions:
# {field_instructions}

# Now, extract the following venue information from the provided text:"""


# # Process venues
# venue_data = []
# test_md_path = Path("test_md")
# if not test_md_path.exists():
#     print(f"Warning: {test_md_path} directory not found")
#     exit()

# md_files = list(test_md_path.glob("*.md"))
# ai_model = "gemini-2.0-flash-exp"  # Adjusted to a valid Gemini model

# for file in tqdm(md_files[2:5], desc="Processing venues", unit="file"):
#     logging.info(f"Processing file: {file.name}")
#     with open(file, "r", encoding="utf-8") as f:
#         md_content = f.read()

#     venue_name = file.stem
#     venue_dict = {"name": venue_name}
#     logging.debug(f"Venue dict initialized for {venue_name}: {venue_dict}")

#     # Process each model for this venue
#     for model_class in models:
#         system_prompt = create_system_prompt(model_class)
#         logging.debug(
#             f"System prompt for {model_class.__name__}: {system_prompt[:500]}..."
#         )  # Limit to first 500 chars
#         try:
#             # Use genai.GenerativeModel directly
#             model = genai.GenerativeModel(model_name=ai_model)
#             logging.debug(f"Generating content with model: {ai_model}")
#             # Skip response_schema for WeddingVenueOther to avoid schema enforcement issues
#             use_schema = model_class != WeddingVenueOther
#             response = model.generate_content(
#                 contents=[
#                     {
#                         "role": "user",
#                         "parts": [
#                             f"{system_prompt}\n\nExtract venue information from this text about '{venue_name}':\n\n{md_content}"
#                         ],
#                     },
#                 ],
#                 generation_config=genai.types.GenerationConfig(
#                     response_mime_type="application/json",
#                     response_schema=model_class if use_schema else None,
#                     temperature=0,
#                 ),
#             )
#             # Extract the raw response correctly
#             if not response.candidates:
#                 raise ValueError(
#                     f"No candidates returned in response for {model_class.__name__}"
#                 )
#             raw_response = response.candidates[0].content.parts[0].text
#             logging.debug(f"Raw response for {model_class.__name__}: {raw_response}")
#             if model_class == WeddingVenueOther:
#                 print(f"Raw response for WeddingVenueOther: {raw_response}")
#             # Attempt to parse JSON
#             try:
#                 venue_info_dict = json.loads(raw_response)
#                 # Convert string numbers to integers for guest_capacity
#                 if (
#                     model_class == WeddingVenueOther
#                     and "guest_capacity" in venue_info_dict
#                 ):
#                     if isinstance(venue_info_dict["guest_capacity"], str):
#                         try:
#                             venue_info_dict["guest_capacity"] = int(
#                                 venue_info_dict["guest_capacity"]
#                             )
#                         except ValueError:
#                             venue_info_dict[
#                                 "guest_capacity"
#                             ] = -1  # Fallback if conversion fails
#             except json.JSONDecodeError as json_error:
#                 logging.error(
#                     f"JSON Decode Error for {model_class.__name__}: {json_error}. Raw response: {raw_response}"
#                 )
#                 print(
#                     f"✗ JSON Decode Error for {model_class.__name__} for {venue_name}: {json_error}. Raw response: {raw_response}"
#                 )
#                 venue_info_dict = {}
#             # Manually validate with Pydantic for WeddingVenueOther
#             if model_class == WeddingVenueOther:
#                 try:
#                     venue_info_dict = model_class(**venue_info_dict).model_dump()
#                 except Exception as pydantic_error:
#                     logging.error(
#                         f"Pydantic Validation Error for {model_class.__name__}: {pydantic_error}. Parsed dict: {venue_info_dict}"
#                     )
#                     print(
#                         f"✗ Pydantic Validation Error for {model_class.__name__} for {venue_name}: {pydantic_error}. Parsed dict: {venue_info_dict}"
#                     )
#                     venue_info_dict = {}
#             # Prefix keys to avoid collisions between models
#             prefixed_venue_info = {
#                 f"{model_class.__name__}_{k}": v for k, v in venue_info_dict.items()
#             }
#             venue_dict.update(prefixed_venue_info)
#             logging.info(
#                 f"✓ Successfully processed {model_class.__name__} for: {venue_name}"
#             )
#         except Exception as e:
#             logging.error(
#                 f"✗ Error processing {model_class.__name__} for {venue_name}: {e}"
#             )
#             print(f"✗ Error processing {model_class.__name__} for {venue_name}: {e}")
#             # Skip WeddingVenueOther with placeholder if error persists
#             if model_class == WeddingVenueOther:
#                 logging.warning(
#                     f"Skipping WeddingVenueOther for {venue_name} due to error: {e}"
#                 )
#                 venue_dict.update({f"{model_class.__name__}_skipped": True})

#     venue_data.append(venue_dict)
#     logging.debug(f"Venue data appended: {venue_dict}")

# # Create and save DataFrame
# if venue_data:
#     df = pd.DataFrame(venue_data)
#     print(f"\nProcessed {len(venue_data)} venues")
#     print(df)

#     now = datetime.now().strftime("%Y%m%d_%H%M%S")
#     output_path = f"venues_data_{now}.csv"
#     df["model"] = ai_model
#     df.to_csv(output_path, index=False)
#     print(f"Data saved to {output_path}")
# else:
#     print("No venue data was processed")

# +
# df_all = pd.concat([df_4o, df_4omini, df_preview, df_gemini])
# df_all = (
#     df_all.drop(columns=["WeddingVenueOther_skipped"])
#     .reset_index()
#     .drop(columns=["index"])
# )

# +
# df_all.to_csv("final_demo_venues_data.csv")

# +
# df

# +
# df.to_excel("final_demo_venues_data_p.xlsx", index=False)
# -
