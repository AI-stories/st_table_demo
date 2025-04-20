import os
from datetime import datetime
from itertools import chain
from pathlib import Path
from typing import Literal

import numpy as np
import openai
import pandas as pd
from google import genai
from openai import OpenAI
from openpyxl.styles import Font, PatternFill
from pydantic import BaseModel, Field
from tqdm import tqdm

TierName = Literal["Standard", "Signature", "Premium"]


class WeddingContactInfo(BaseModel):
    city: str | None = Field(description="The city this wedding venue is located in.")
    state: str | None = Field(description="The state this wedding venue is located in.")
    zip_code: str | None = Field(
        description="The zip code this wedding venue is located in."
    )
    country: str | None = Field(
        description="The country this wedding venue is located in."
    )
    email: str | None = Field(description="The email address of the wedding venue.")
    phone: str | None = Field(description="The phone number of the wedding venue.")
    website: str | None = Field(description="The website of the wedding venue.")
    facebook: str | None = Field(description="The facebook page of the wedding venue.")
    instagram: str | None = Field(
        description="The instagram account of the wedding venue."
    )


class MenuTier(BaseModel):
    name: TierName = Field(
        description="Tier name: must be Standard, Signature, or Premium."
    )
    appetizers: str | None = Field(description="Appetizers in this tier.")
    entrees: str | None = Field(description="Entrées in this tier.")
    sides: str | None = Field(description="Side dishes in this tier.")
    desserts: str | None = Field(description="Desserts in this tier.")
    beverages: str | None = Field(description="Beverages in this tier.")

    def to_string(self) -> str:
        parts = [f"{self.name}"]
        if self.appetizers:
            parts.append(f"- Appetizers: {self.appetizers}")
        if self.entrees:
            parts.append(f"- Entrées: {self.entrees}")
        if self.sides:
            parts.append(f"- Sides: {self.sides}")
        if self.desserts:
            parts.append(f"- Desserts: {self.desserts}")
        if self.beverages:
            parts.append(f"- Beverages: {self.beverages}")
        if len(parts) == 1:
            return ""
        return "\n".join(parts)


class FoodBreakdown(BaseModel):
    """
    You must return exactly three food tiers: Standard, Signature, and
    Premium.

    Venues often name their packages arbitrarily (e.g., "Adore", "Silver",
    "Classic", "Treasure"). Your task is to interpret these names and map each one
    to a standard tier by evaluating the **menu quality, service level, and
    pricing**.

    Use the following general guidance:

    - **Standard**: basic offerings, lower price point
    - **Signature**: moderate enhancements, mid-range price
    - **Premium**: multiple high-end appetizers, upgraded entrées (beef,
        seafood), top-tier service, highest price

    Ignore the original package names. Focus only on the **content and price range**
    to assign the correct standard tier name.
    """

    tiers: list[MenuTier] = Field(
        description="List of exactly 3 tiers: Standard (Silver), Signature (Gold), Premium (Platinum)."
    )
    flexibility: Literal[
        "0: Completely fixed packages, no flexibility",
        "1: Fixed packages with a few extras or options",
        "2: Moderate or flexible approach",
        "3: Highly customizable with some structure",
        "4: Completely custom/DIY",
    ] = Field(
        description="How much freedom does the customer have to customize the food-related tier packages?"
    )

    def to_string(self) -> str:
        return "\n\n".join([tier.to_string() for tier in self.tiers])


class BarTier(BaseModel):
    name: TierName = Field(
        description="Tier name: must be Standard, Signature, or Premium."
    )
    highlights: str | None = Field(
        description="Beverages and package highlights in this drink package."
    )

    bar_pricing_model: Literal["Open bar", "Hosted bar", "Cash bar", "Not Offered"] = (
        Field(
            description="""- open bar: the venue provides the bar and the drinks, prepaid for by the host
- hosted bar: the venue provides the bar and the drinks, paid for by the host at the end of the night, also referred to as a consumption bar or tab bar
- cash bar: the venue provides the bar, but the drinks are payed for by the guests
- not offered: the venue does not offer a bar or is not specified"""
        )
    )

    def to_string(self) -> str:
        parts = [f"{self.name}"]
        if self.highlights:
            parts.append(f"- Highlights: {self.highlights}")
        if self.bar_pricing_model != "Not Offered":
            parts.append(f"- Bar Pricing Model: {self.bar_pricing_model}")
        return "\n".join(parts)


class BarBreakdown(BaseModel):
    """
    You must return exactly three bar tiers: Standard, Signature, and
    Premium.

    Many venues use unique names for their drink packages. Instead of
    copying those names, evaluate the **included alcohol types and
    service level**, and map them to a standard tier:

    - **Standard**: basic offerings, lower price point
    - **Signature**: mid-range price, includes house liquors or a soft
        bar with wine service
    - **Premium**: top-tier service, highest price like top-shelf
        liquor, signature cocktails, champagne toast, or full open bar

    Always normalize to these three tiers based on the drink offerings —
    not the label.
    """

    tiers: list[BarTier] = Field(
        description="List of exactly 3 tiers: Standard (Silver), Signature (Gold), Premium (Platinum)."
    )
    flexibility: Literal[
        "0: Completely fixed packages, no flexibility",
        "1: Fixed packages with a few extras or options",
        "2: Moderate or flexible approach",
        "3: Highly customizable with some structure",
        "4: Completely custom/DIY",
    ] = Field(
        description="How much freedom does the customer have to customize the package?"
    )

    def to_string(self) -> str:
        parts = []
        for tier in self.tiers:
            parts.append(tier.to_string())

        return "\n".join(parts)


class WeddingPriceInfo(BaseModel):
    option: Literal["standard", "premium", "signature"] = Field(
        description="This is the pricing option for this wedding venue."
    )
    ceremony_cost: int | Literal["unknown"] = Field(
        description="Estimate a cost for the ceremony. Respond with a number only which would be your best guess based on the available information. Use 'unknown' if no information is provided."
    )

    def to_string(self) -> str:
        if self.ceremony_cost:
            return f"Ceremony: {self.option}\n - ceremony cost: ${self.ceremony_cost}\n"
        return ""


class WeddingVenueStyle(BaseModel):
    style: Literal[
        "Barns & Farms",
        "Hotels",
        "Winery",
        "Country Clubs",
        "Restaurants",
        "Rooftops & Lofts",
        "Mansions",
        "Religious Spaces",
        "Museums",
        "Boats",
        "Parks",
        "Historic Venues",
        "Banquet Halls",
        "Beach",
        "Garden",
        "Waterfront",
        "Brewery",
        "State",
        "Local",
        "Government Property",
        "Other",
    ] = Field(
        description="""Determine the style of the venue. You MUST CHOOSE ONE of the following options that best matches the document's content and return the selected option's value as the field value:
- Barns & Farms
- Hotels
- Winery
- Country Clubs
- Restaurants
- Rooftops & Lofts
- Mansions
- Religious Spaces
- Museums
- Boats
- Parks
- Historic Venues
- Banquet Halls
- Beach
- Garden
- Waterfront
- Brewery
- State
- Local
- Government Property
- Other"""
    )
    indoor_outdoor: Literal[
        "1: Completely indoor",
        "2: Predominantly indoor",
        "3: Equal indoor/outdoor mix",
        "4: Predominantly outdoor",
        "5: Completely outdoor",
        "X: Not enough information",
    ] = Field(
        description="""Determine the indoor/outdoor nature of the venue. You MUST CHOOSE ONE of the following options and return the selected option's description as the field value:
- "1: Completely indoor"
- "2: Predominantly indoor"
- "3: Equal indoor/outdoor mix"
- "4: Predominantly outdoor"
- "5: Completely outdoor"
- "X: Not enough information"

For guidance: You must choose one of the above options. A predominantly outdoor space may include gardens or patios, while one without a covering lacks weather protection (e.g., an open field)."""
    )
    privacy: Literal[
        "1: Privacy and exclusivity is a major feature of the venue",
        "2: Moderate privacy with possible nearby non-wedding guests",
        "3: Shared or public space",
        "X: Not enough information",
    ] = Field(
        description="""Assess the privacy level. You MUST CHOOSE ONE of the following options and return the selected option's description as the field value:
- "1: Privacy and exclusivity is a major feature of the venue"
- "2: Moderate privacy with possible nearby non-wedding guests"
- "3: Shared or public space"
- "X: Not enough information"."""
    )
    accommodations: Literal[
        "1: On-site lodging accommodations or extremely close (less than 5 mins walking distance) is possible",
        "2: No On-site lodging, third-party 5-10 minutes away.",
        "3: No on-site lodging, third-party 10+ minutes away.",
        "X: Not enough information",
    ] = Field(
        description="""Determine the lodging accommodations offered by the venue. You MUST CHOOSE ONE of the following options that best matches the document's content and return the selected option's description as the field value:
- "1: On-site lodging accommodations or extremely close (less than 5 mins walking distance) is possible"
- "2: No On-site lodging, third-party 5-10 minutes away."
- "3: No on-site lodging, third-party 10+ minutes away."
- "X: Not enough information"."""
    )
    environmental: Literal[
        "This venue focuses on minimal environmental impact and sustainability in their offering",
        "This venue does not emphasize environment or sustainability",
    ] = Field(
        description="""Assess the venue's emphasis on environmental sustainability. You MUST CHOOSE ONE of the following options that best matches the document's content and return the selected option's description as the field value:
- This venue focuses on minimal environmental impact and sustainability in their offering
- This venue does not emphasize environment or sustainability.

For guidance: A focus on sustainability might include eco-friendly practices (e.g., local sourcing or renewable energy), while lack of emphasis implies no mention of such efforts."""
    )
    general_vibe: Literal[
        "Rustic and simple",
        "Peaceful and serene",
        "Grandiose and elegant",
        "Adventurous or quirky",
        "Warm and cozy",
        "Other",
    ] = Field(
        description="""Determine the general vibe or atmosphere of the venue. You MUST CHOOSE ONE of the following options that best matches the document's content and return the selected option's description as the field value:
- Rustic and simple
- Peaceful and serene
- Grandiose and elegant
- Adventurous or quirky
- Warm and cozy
- Other. For guidance: Rustic and simple might include barns or farms, peaceful and serene could be gardens, grandiose and elegant might be mansions, adventurous or quirky could involve unique spaces, and warm and cozy might be cozy interiors."""
    )


class WeddingVenueOther(BaseModel):
    outside_wedding_coordinator: Literal["True", "False", "unknown"] = Field(
        description="This venue lets me bring in my own wedding coordinator"
    )
    outside_photographer: Literal["True", "False", "unknown"] = Field(
        description="This venue lets me bring in my own photographer"
    )
    package_approach: Literal[
        "1: This venue offers fixed packages, with a few extras or options",
        "2: This venue offers a moderate or flexible approach",
        "3: This venue provides a high degree of flexibility",
        "X: Not enough information",
    ] = Field(
        description="""Determine the wedding package approach of the venue. You MUST CHOOSE ONE of the following options that best matches the document's content and return the selected option's description as the field value:
- "1: This venue offers fixed packages, with a few extras or options"
- "2: This venue offers a moderate or flexible approach"
- "3: This venue provides a high degree of flexibility"
- "X: Not enough information"

For guidance: Fixed packages include in-house services (e.g., pre-set menus); moderate offers a mix; high flexibility requires client planning."""
    )
    reception_or_ceremony: Literal[
        "1: Only space for reception",
        "2: Only space for the ceremony",
        "3: Space for both reception and the ceremony",
        "X: Not enough information",
    ] = Field(
        description="""Determine an option for spaces provided by the venue. You MUST CHOOSE ONE of the following options that best matches the document's content and return the selected option's description as the field value:
- "1: Only space for reception"
- "2: Only space for the ceremony"
- "3: Space for both reception and the ceremony"
- "X: Not enough information"."""
    )
    what_time_does_the_party_need_to_stop: Literal[
        "1: 10PM",
        "2: 12AM",
        "3: After 12AM",
        "X: Not enough information",
    ] = Field(
        description="""Determine the latest time the event must conclude. You MUST CHOOSE ONE of the following options that best matches the document's content:
- "1: 10PM"
- "2: 12AM"
- "3: After 12AM"
- "X: Not enough information"."""
    )
    top_choices: str = Field(
        description="Give me 1-2 reasons why you think a client is most likely to choose this venue (e.g., what are the biggest key selling points or benefits)."
    )
    guest_capacity: Literal["1", "50", "100", "150", "200", "300"] = Field(
        description="""Determine the estimated guest capacity. You MUST CHOOSE ONE of the following approximate maximums that best matches the document's content, representing the upper bound of each range:
- 1 (for 1-50)
- 50 (for 50-100)
- 100 (for 100-150)
- 150 (for 150-200)
- 200 (for 200-300)
- 300 (for 300+)
Guidance: Infer from max guests or room size (e.g., 'up to 150 guests' = 150)."""
    )


def generate_field_instructions(model_class: type[BaseModel]) -> str:
    """Generate field-specific instructions from a Pydantic model's docstrings."""
    instructions = []
    for field_name, field_info in model_class.model_fields.items():
        docstring = (
            field_info.description or field_info.__doc__
        ) or "No description provided."
        instructions.append(f"- {field_name}: {docstring}")
    return "\n".join(instructions)


def create_system_prompt(model_class: type[BaseModel]) -> str:
    field_instructions = generate_field_instructions(model_class)
    tier_hint = model_class.__doc__

    return f"""
        You are an expert in wedding planning. You are extracting structured
        information about wedding venues.

        First, carefully analyze all relevant information in the text. Consider
        both explicit statements and reasonable inferences.

        Important instructions: 
        1. For each field, follow the specific guidelines below about how to handle 
           ambiguous or missing information.
        2. For boolean fields, return true/false values rather than "Yes"/"No" strings. 
        3. For string fields, provide detailed information or null if not available. 
        4. Begin by developing a comprehensive reasoning that considers all evidence 
           before determining individual field values.

        Field-specific instructions:
        {tier_hint}
        {field_instructions}
        """


class WeddingFoodInfo(BaseModel):
    outside_food_allowed: bool = Field(
        description="""Can external catering or food vendors be used at this venue?If the text explicitly mentions allowing outside catering or food vendors, answer True. Do not consider desserts or cakes as food, as these are different categories. If not mentioned, default to False."""
    )
    outside_alcohol_allowed: bool = Field(
        description="""Can external alcohol be brought into this venue?If the text mentions the venue has a liquor license or provides alcohol service, answer False. If the text explicitly mentions allowing outside alcohol or BYOB policy, answer True. If not mentioned, default to False."""
    )
    outside_dessert_allowed: bool = Field(
        description="""Can external desserts or cakes be brought into this venue?If the text explicitly mentions allowing outside desserts or cakes, answer True. If not mentioned, default to True."""
    )
    kosher_food: bool = Field(
        description="""Does the venue offer certified kosher menu options, either in-house or through external vendors?Answer True ONLY if kosher options are explicitly mentioned or if the venue explicitly states they accommodate religious dietary restrictions. If not mentioned, default to False."""
    )
    halal_food: bool = Field(
        description="""Does the venue offer certified halal menu options, either in-house or through external vendors? Return True if offered, False if not offered. Answer True ONLY if halal options are explicitly mentioned. If not mentioned, default to False."""
    )
    east_asian_food: bool = Field(
        description="""Does the venue offer East Asian food options (Chinese, Japanese, Korean, etc.), either in-house or through external vendors? Answer True if East Asian cuisine is explicitly mentioned or if the venue offers international cuisine. If not mentioned, default to False."""
    )
    indian_food: bool = Field(
        description="""Does the venue offer Indian food options, either in-house or through external vendors? Answer True if Indian cuisine is explicitly mentioned or if the venue offers diverse international cuisine that likely includes Indian options. If not mentioned, default to False."""
    )
    gluten_free_food: bool = Field(
        description="""Does the venue offer gluten-free food options? Return True if offered, False if not offered. Answer True if gluten-free options are explicitly mentioned. If the venue mentions accommodating dietary restrictions/allergies/food preferences, return True."""
    )
    other_ethnic_food_style: str = Field(
        description="""Does the venue offer other ethnic food styles beyond those already mentioned? If so, list the available styles as a comma-separated string. If no other ethnic food styles are mentioned, return 'unknown'."""
    )
    late_night_food: bool = Field(
        description="""Does the venue provide late-night food options, such as pizza or snacks after dinner and dessert service? If not mentioned, default to False."""
    )


class WeddingVenuePricingSummary(BaseModel):
    # options: list[WeddingPriceInfo] = Field(
    #     description="This is a list of all the pricing options for this wedding venue."
    # )
    price: int = Field(
        description="""Analyze this wedding venue document and calculate the PER PERSON cost in USD for a 100-guest wedding reception.

Important instructions:
1. Return ONLY a single number representing the per-person cost in USD (e.g., 250)
2. Do NOT return the total venue cost - I need the PER PERSON amount
3. If you see a total price (e.g., $20,000 for 80 guests), divide by the number of guests to get the per-person rate ($20,000 ÷ 80 = $250 per person)
4. Then adjust this per-person rate for a 100-person event, accounting for any tiered pricing

Your calculation must include:
- Food (dinner, appetizer, dessert)
- Beverage/alcohol package
- Venue rental fees (divided by 100 guests)
- All taxes, service charges, and gratuities

Exclude:
- Ceremony costs (focus only on reception)
- Photography
- Late night food
- Any optional add-ons

If multiple packages exist, use the middle-tier or "standard" option.
If specific costs aren't provided, estimate based on the venue's location and quality level.

Think through your calculation step by step:
1. Identify the base per-person food & beverage cost
2. Add any flat venue fees (calculate the per-person amount)
3. Add per-person share of taxes and service charges
4. Sum these components for the final per-person total

Remember: Your answer must be ONLY a single number representing the per-person cost in USD."""
    )
    base_prices: str = Field(
        description="""The base cost of the venue, including menu items,
                    beverages and any variations based on day of the week or
                    package. Break down ALL venue costs on a per-person basis.
                    For example, if the venue has a $20,000 base price for 80
                    guests, the base price per person is $250. If there is not
                    enough information, take the base assumption of $100 per
                    person for food and $50 per person for beverage."""
    )
    taxes_and_fees: str = Field(
        description="The breakdown of taxes, gratuity, and any additional fees, including their individual amounts and the total. Convert ALL taxes, service charges, and fees to per-person amounts. For example, if the venue has a $2,185 tax for 80 guests, the tax per person is $27.31."
    )
    pricing_transparency: Literal[
        "1: This venue discloses a small portion of the total wedding costs",
        "2: This venue discloses a moderate portion of the total wedding costs",
        "3: This venue discloses a high degree of the total costs",
        "X: Not enough information",
    ] = Field(
        description="""Assess how much of the total wedding cost is disclosed in the provided materials. You MUST CHOOSE ONE of the following options that best matches the document's content and return the selected option's description as the field value:

- This venue discloses a small portion of the total wedding costs
- This venue discloses a moderate portion of the total wedding costs
- This venue discloses a high degree of the total wedding costs
- Not enough information

Guidance when selecting the option: A 'small portion' of disclosure means significant costs (e.g., food, bar/alcohol) are unclear or require contacting external vendors. A 'moderate portion' means some unknowns exist, but you can get a general cost idea without much extra work. A 'high degree' means most costs are disclosed with few surprises, little additional work needed to understand the total cost."""
    )
    flexibility: Literal[
        "0: Completely fixed packages, no flexibility",
        "1: Fixed packages with a few extras or options",
        "2: Moderate or flexible approach",
        "3: Highly customizable with some structure",
        "4: Completely custom/DIY",
    ] = Field(
        description="How much freedom does the customer have to customize the wedding/ceremony package?"
    )
    deposit_and_payment_plans: Literal[
        "The venue works with me on deposit terms and payment plans",
        "The venue does not have flexibility on deposit terms and payment plans",
        "Not enough information",
    ] = Field(
        description="""Determine if the venue offers flexibility on deposit terms and payment plans. You MUST CHOOSE ONE of the following options that best matches the document's content and return the selected option's description as the field value:

- The venue works with me on deposit terms and payment plans
- The venue does not have flexibility on deposit terms and payment plans
- Not enough information.

Follow these guidance when selecting the option: Flexibility means the venue allows negotiation on deposit amounts, payment schedules, or offers installment plans. Lack of flexibility is indicated by strict terms or no mention of flexible options."""
    )


models = [
    WeddingContactInfo,
    FoodBreakdown,
    WeddingFoodInfo,
    BarBreakdown,
    WeddingVenuePricingSummary,
    WeddingVenueStyle,
    WeddingVenueOther,
]


def assert_keys_in_readable_columns(
    models: list[type[BaseModel]], readable_columns: dict[str, str]
) -> None:
    keys = list(
        chain.from_iterable(
            [model.__name__ + "_" + x for x in model.model_fields.keys()]
            for model in models
        )
    )
    keys = [
        str(key).replace("_tiers", "_summary").replace("_options", "_summary")
        for key in keys
    ]
    keys
    assert set(keys) - set(readable_columns.keys()) == set(), (
        f"missing keys in readable_columns: {set(keys) - set(readable_columns.keys())}"
    )


readable_columns = {
    "venue": "wedding venue",
    "WeddingVenuePricingSummary_price": "price per guest",
    "WeddingVenuePricingSummary_base_prices": "price breakdown",
    "WeddingVenuePricingSummary_taxes_and_fees": "price breakdown taxes and fees",
    "WeddingVenuePricingSummary_flexibility": "venue customization flexibility",
    # "WeddingPriceInfo_option": "options",
    "WeddingContactInfo_city": "city",
    "WeddingContactInfo_state": "state",
    "WeddingContactInfo_country": "country",
    "WeddingContactInfo_zip_code": "zip code",
    "WeddingContactInfo_email": "email",
    "WeddingContactInfo_website": "website",
    "WeddingContactInfo_phone": "phone",
    "WeddingContactInfo_facebook": "facebook",
    "WeddingContactInfo_instagram": "instagram",
    # "WeddingVenuePricingSummary_summary": "venue pricing summary",
    "FoodBreakdown_summary": "food menu breakdown",
    "FoodBreakdown_flexibility": "food menu flexibility",
    "BarBreakdown_summary": "bar menu breakdown",
    "BarBreakdown_flexibility": "bar menu flexibility",
    "WeddingVenuePricingSummary_pricing_transparency": "pricing transparency",
    "WeddingVenuePricingSummary_deposit_and_payment_plans": "deposit and payment plans",
    "WeddingVenueStyle_style": "style",
    "WeddingVenueStyle_indoor_outdoor": "indoor/outdoor seating",
    "WeddingVenueStyle_privacy": "privacy",
    "WeddingVenueStyle_accommodations": "accommodations",
    "WeddingVenueStyle_environmental": "environmental",
    "WeddingVenueStyle_general_vibe": "general vibe",
    "WeddingFoodInfo_east_asian_food": "serves east asian food",
    "WeddingFoodInfo_gluten_free_food": "serves gluten free food",
    "WeddingFoodInfo_halal_food": "serves halal food",
    "WeddingFoodInfo_indian_food": "serves indian food",
    "WeddingFoodInfo_kosher_food": "serves kosher food",
    "WeddingFoodInfo_late_night_food": "serves late night food",
    "WeddingFoodInfo_other_ethnic_food_style": "serves other ethnic food",
    "WeddingFoodInfo_outside_alcohol_allowed": "allows outside alcohol",
    "WeddingFoodInfo_outside_dessert_allowed": "allows outside dessert",
    "WeddingFoodInfo_outside_food_allowed": "allows outside food",
    "WeddingVenueOther_guest_capacity": "guest capacity",
    "WeddingVenueOther_what_time_does_the_party_need_to_stop": "what time does the party need to stop",
    "WeddingVenueOther_outside_photographer": "allows outside photographer",
    "WeddingVenueOther_package_approach": "package approach",
    "WeddingVenueOther_outside_wedding_coordinator": "allows outside wedding coordinator",
    "WeddingVenueOther_reception_or_ceremony": "reception or ceremony",
    "WeddingVenueOther_top_choices": "top choices",
}


assert_keys_in_readable_columns(models, readable_columns)


def flatten_dict(d: dict, parent_key: str = "", sep: str = "_") -> dict:
    """Flatten a nested dictionary by concatenating nested keys with a separator.

    Parameters
    ----------
    d : dict
        The dictionary to flatten
    parent_key : str, optional
        The parent key for nested dictionaries, by default ""
    sep : str, optional
        The separator to use between nested keys, by default "_"

    Returns
    -------
    dict
        A flattened dictionary with concatenated keys

    Examples
    --------
    >>> d = {"a": 1, "b": {"c": 2, "d": {"e": 3}}}
    >>> flatten_dict(d)
    {'a': 1, 'b_c': 2, 'b_d_e': 3}
    """
    items: list = []
    for k, v in d.items():
        new_key = f"{parent_key}{sep}{k}" if parent_key else k
        if isinstance(v, dict):
            items.extend(flatten_dict(v, new_key, sep).items())
        else:
            items.append((new_key, v))
    return dict(items)


class WeddingVenue:
    def __init__(self, venue_name: str, raw: list[BaseModel]):
        item_dict = {"venue": venue_name}
        for item in raw:
            obj_dict = item.model_dump()
            if "tiers" in obj_dict:
                obj_dict.pop("tiers")
                obj_dict["summary"] = item.to_string()

            if "options" in obj_dict:
                obj_dict.pop("options")
                obj_dict["summary"] = item.to_string()

            item_dict[item.__class__.__name__] = obj_dict

        self.df = pd.DataFrame()
        self.update(item_dict)

    def add_price_breakdown(self) -> None:
        self.df["price breakdown"] = self.df[
            [
                "price breakdown",
                "price breakdown taxes and fees",
            ]
        ].apply(
            lambda x: f"""
                base prices: {x.iloc[0]}
                taxes and fees: {x.iloc[1]}
                """,
            axis=1,
        )
        del self.df["price breakdown taxes and fees"]

    def update(self, d: dict) -> None:
        self.df = pd.DataFrame(flatten_dict(d), index=[0])
        self.rename_columns()
        self.add_price_breakdown()
        # self.add_bar_flexibility()
        # self.add_indoor_outdoor_seating()

    def _repr_html_(self) -> str:
        return self.df._repr_html_()

    def rename_columns(self) -> None:
        """Rename and reorder columns based on readable_columns dictionary."""
        self.df.rename(columns=readable_columns, inplace=True)
        self.df.set_index("wedding venue", inplace=True)
        ordered_columns = [
            col for col in readable_columns.values() if col != "wedding venue"
        ]
        self.df = self.df.reindex(columns=ordered_columns)

    def to_excel(self, name: str = "wedding_venue.xlsx"):
        if not name.endswith(".xlsx"):
            name = f"{name}.xlsx"
        if os.path.exists(name):
            name = f"{name}_{datetime.now().strftime('%Y%m%d_%H%M%S')}.xlsx"
        print("saving to: ", name)
        with pd.ExcelWriter(name, engine="openpyxl") as writer:
            self.df.to_excel(writer, sheet_name="Venue Options")

            worksheet = writer.sheets["Venue Options"]

            header_fill = PatternFill(
                start_color="B3E5FC", end_color="B3E5FC", fill_type="solid"
            )
            header_font = Font(bold=True)

            for col in range(1, len(self.df.columns) + 1):
                cell = worksheet.cell(row=1, column=col)
                cell.fill = header_fill
                cell.font = header_font

            for col in worksheet.columns:
                max_length = 0
                column = col[0].column_letter
                for cell in col:
                    if cell.value:
                        max_length = max(max_length, len(str(cell.value)))
                adjusted_width = max_length + 2
                worksheet.column_dimensions[column].width = min(adjusted_width, 50)

            worksheet.auto_filter.ref = worksheet.dimensions

        self.df.to_excel(writer, sheet_name="Venue Options")
        return self

    def __add__(self, other: "WeddingVenue") -> "WeddingVenue":
        self.df = pd.concat([self.df, other.df])
        return self


class Response:
    def __init__(self, ai: Literal["openai", "google"]):
        self.ai = ai
        if ai == "openai":
            self.client = OpenAI(api_key=os.getenv("OPENAI_API_KEY"))
            self.response = self._response_openai
        if ai == "google":
            self.client = genai.Client(api_key=os.getenv("GEMINI_API_KEY"))
            self.response = self._response_google

    def _response_openai(
        self, model, system_prompt, user_prompt, response_format, temperature
    ):
        completion = self.client.beta.chat.completions.parse(
            model=model,
            messages=[
                {"role": "system", "content": system_prompt},
                {
                    "role": "user",
                    "content": user_prompt,
                },
            ],
            response_format=response_format,
            temperature=temperature,
        )
        return completion.choices[0].message.parsed

    def _response_google(
        self, model, system_prompt, user_prompt, response_format, temperature
    ):
        response = self.client.models.generate_content(
            model=model,
            contents=f"{system_prompt}\n{user_prompt}",
            config={
                "response_mime_type": "application/json",
                "response_schema": response_format,
            },
        )
        return response.parsed


def markdown_to_structured_output(
    markdown_path: str | Path, ai: Literal["openai", "google"] = "openai"
) -> WeddingVenue:
    response = Response(ai=ai)
    venue = markdown_path.parent.name
    text_content = Path(markdown_path).read_text()
    raw_output = []
    venue_dict = {"name": venue}
    for model_class in models:
        try:
            system_prompt = create_system_prompt(model_class)
            if response.ai == "openai":
                if model_class == WeddingVenuePricingSummary:
                    ai_model = "o3-mini"
                    temperature = openai.NOT_GIVEN
                else:
                    ai_model = "gpt-4o-mini"
                    temperature = 0
                obj = response.response(
                    model=ai_model,
                    system_prompt=system_prompt,
                    user_prompt=f"Extract venue information from this text about '{venue}':\n\n{text_content}",
                    response_format=model_class,
                    temperature=temperature,
                )
            elif response.ai == "google":
                temperature = 0
                ai_model = "gemini-2.0-flash-001"
                obj = response.response(
                    model=ai_model,
                    system_prompt=system_prompt,
                    user_prompt=f"Extract venue information from this text about '{venue}':\n\n{text_content}",
                    response_format=model_class,
                    temperature=temperature,
                )
            raw_output.append(obj)
            if hasattr(obj, "to_string"):
                string_summary = obj.to_string()
                venue_dict[f"{model_class.__name__}_summary"] = string_summary
            else:
                venue_dict[f"{model_class.__name__}_summary"] = obj.model_dump()

            tqdm.write(f"✓ Processed {model_class.__name__} for: {venue}")

        except Exception as e:
            tqdm.write(f"✗ Error with {model_class.__name__} for {venue}: {e}")
            venue_dict[f"{model_class.__name__}_summary"] = None

    return WeddingVenue(venue, raw_output)
